#!/usr/bin/env python3
"""DORA Metrics CLI — user interaction and orchestration."""

import argparse
import asyncio
import json
import sys
from datetime import UTC, datetime
from pathlib import Path

import httpx
import openpyxl

from azure_api import (
    fetch_all_builds_for_project,
    fetch_builds,
    fetch_pipelines,
    fetch_projects,
    fetch_pull_requests,
    fetch_repos,
    parse_org,
    read_pat,
)
from dora_metrics import (
    all_months_in_range,
    classify_dora,
    compute_change_failure_rate_by_month,
    compute_deployment_frequency,
    compute_lead_times_by_month,
    compute_mttr_by_month,
    compute_pr_deployment_frequency,
    compute_pr_lead_times_by_month,
    format_hours,
)


def prompt_choice(items: list[str], label: str, allow_all: bool = False) -> int | None:
    """Display numbered list, return selected index or None for 'all'."""
    for i, item in enumerate(items, 1):
        print(f"  {i}. {item}")
    if allow_all:
        print(f"  0. (All)")
    while True:
        raw = input(f"\nSelect {label}: ").strip()
        if not raw.isdigit():
            print("Please enter a number.")
            continue
        n = int(raw)
        if allow_all and n == 0:
            return None
        if 1 <= n <= len(items):
            return n - 1
        print(f"Please enter a number between {'0' if allow_all else '1'} and {len(items)}.")


def print_results(df: dict, lt: dict, cfr: dict, mttr: dict, months: list[str], title: str = ""):
    col_w = 14
    label_w = 28
    cols = months + ["OVERALL"]
    w = label_w + col_w * len(cols) + 2
    sep = "─" * w

    header = title if title else "DORA METRICS REPORT"
    report_date = datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")

    print("\n" + "=" * w)
    print(f"  {header}")
    print(f"  Generated: {report_date}")
    print("=" * w)

    # Header row
    hdr = f"  {'Metric':<{label_w}}"
    for c in cols:
        hdr += f"{c:>{col_w}}"
    print(f"\n{sep}")
    print(hdr)
    print(sep)

    # Helper to get overall values
    lt_overall = lt.get("_overall", {})
    cfr_overall = cfr.get("_overall", {})
    mttr_overall = mttr.get("_overall", {})

    # Row 1: Deployment Frequency (days/deploy)
    row = f"  {'Deploy Freq (days/dep)':<{label_w}}"
    for mk in months:
        df_m = df["monthly"].get(mk)
        dpd = df_m["days_per_dep"] if df_m else None
        val = f"{dpd:.1f}" if dpd is not None else "N/A"
        row += f"{val:>{col_w}}"
    ov = f"{df['overall_days_per_dep']:.1f}" if df["overall_days_per_dep"] is not None else "N/A"
    row += f"{ov}".rjust(col_w)
    print(row)

    # Row 2: Deploy count
    row = f"  {'  (total deploys)':<{label_w}}"
    for mk in months:
        df_m = df["monthly"].get(mk)
        val = str(df_m["count"]) if df_m else "0"
        row += f"{val:>{col_w}}"
    row += f"{df['total']}".rjust(col_w)
    print(row)

    print(sep)

    # Row 3: Lead Time for Changes
    row = f"  {'Lead Time for Changes':<{label_w}}"
    for mk in months:
        lt_m = lt.get(mk)
        val = format_hours(lt_m["avg_hours"]) if lt_m else "N/A"
        row += f"{val:>{col_w}}"
    row += f"{format_hours(lt_overall.get('avg_hours'))}".rjust(col_w)
    print(row)

    # Row 4: Lead time sample size
    row = f"  {'  (sample size)':<{label_w}}"
    for mk in months:
        lt_m = lt.get(mk)
        val = str(lt_m["sample_size"]) if lt_m else "0"
        row += f"{val:>{col_w}}"
    row += f"{lt_overall.get('sample_size', 0)}".rjust(col_w)
    print(row)

    print(sep)

    # Row 5: Change Failure Rate
    row = f"  {'Change Failure Rate':<{label_w}}"
    for mk in months:
        cfr_m = cfr.get(mk)
        rate = cfr_m["rate_pct"] if cfr_m else None
        val = f"{rate:.1f}%" if rate is not None else "N/A"
        row += f"{val:>{col_w}}"
    cfr_ov = f"{cfr_overall['rate_pct']:.1f}%" if cfr_overall.get("rate_pct") is not None else "N/A"
    row += f"{cfr_ov}".rjust(col_w)
    print(row)

    # Row 6: CFR detail
    row = f"  {'  (failed / total)':<{label_w}}"
    for mk in months:
        cfr_m = cfr.get(mk, {"failed": 0, "total": 0})
        val = f"{cfr_m['failed']}/{cfr_m['total']}"
        row += f"{val:>{col_w}}"
    row += f"{cfr_overall.get('failed', 0)}/{cfr_overall.get('total', 0)}".rjust(col_w)
    print(row)

    print(sep)

    # Row 7: MTTR
    row = f"  {'MTTR':<{label_w}}"
    for mk in months:
        mttr_m = mttr.get(mk)
        val = format_hours(mttr_m["avg_hours"]) if mttr_m else "N/A"
        row += f"{val:>{col_w}}"
    row += f"{format_hours(mttr_overall.get('avg_hours'))}".rjust(col_w)
    print(row)

    # Row 8: MTTR incidents
    row = f"  {'  (incidents)':<{label_w}}"
    for mk in months:
        mttr_m = mttr.get(mk)
        val = str(mttr_m["incidents"]) if mttr_m else "0"
        row += f"{val:>{col_w}}"
    row += f"{mttr_overall.get('incidents', 0)}".rjust(col_w)
    print(row)

    print(sep)

    # DORA categories row
    print(f"\n{sep}")
    print("  DORA PERFORMANCE CATEGORIES")
    print(sep)

    for metric_key, metric_label, get_val in [
        ("deploy_freq", "Deploy Frequency", lambda mk: (df["monthly"].get(mk, {}) or {}).get("days_per_dep")),
        ("lead_time", "Lead Time", lambda mk: (lt.get(mk) or {}).get("avg_hours")),
        ("cfr", "Change Failure Rate", lambda mk: (cfr.get(mk) or {}).get("rate_pct")),
        ("mttr", "MTTR", lambda mk: (mttr.get(mk) or {}).get("avg_hours")),
    ]:
        row = f"  {metric_label:<{label_w}}"
        for mk in months:
            val = get_val(mk)
            cat = classify_dora(metric_key, val)
            row += f"{cat:>{col_w}}"
        # Overall
        overall_values = {
            "deploy_freq": df["overall_days_per_dep"],
            "lead_time": lt_overall.get("avg_hours"),
            "cfr": cfr_overall.get("rate_pct"),
            "mttr": mttr_overall.get("avg_hours"),
        }
        row += f"{classify_dora(metric_key, overall_values[metric_key])}".rjust(col_w)
        print(row)

    print(f"\n{'=' * w}")
    print("  Legend: Elite > High > Medium > Low")
    print("=" * w)


def export_json(
    org: str, project: str, mode: str, months: list[str],
    df: dict, lt: dict, cfr: dict, mttr: dict,
) -> str:
    """Export metrics to a JSON file under reports/. Returns the filepath."""
    def strip_details(d: dict) -> dict:
        return {k: v for k, v in d.items() if k != "_details"}

    def get_details(d: dict) -> dict:
        return d.get("_details", {})

    payload = {
        "generated": datetime.now(UTC).isoformat(),
        "organization": org,
        "project": project,
        "mode": mode,
        "months": months,
        "summary": {
            "deployment_frequency": strip_details(df),
            "lead_time": strip_details(lt),
            "change_failure_rate": strip_details(cfr),
            "mttr": strip_details(mttr),
        },
        "details": {
            "deployment_frequency": get_details(df),
            "lead_time": get_details(lt),
            "change_failure_rate": get_details(cfr),
            "mttr": get_details(mttr),
        },
    }

    reports_dir = Path(__file__).resolve().parent / "reports"
    reports_dir.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_project = project.replace(" ", "_")
    filepath = reports_dir / f"dora_{safe_project}_{mode}_{timestamp}.json"
    with open(filepath, "w") as f:
        json.dump(payload, f, indent=2, default=str)
    print(f"\nJSON report exported to: {filepath}")
    return str(filepath)


def export_excel(
    org: str, project: str, mode: str, months: list[str],
    df: dict, lt: dict, cfr: dict, mttr: dict,
) -> str:
    """Export metrics to an Excel copy of DORA_DB.xlsx. Returns the filepath."""
    from azure_api import parse_dt

    template = Path(__file__).resolve().parent / "DORA_DB.xlsx"
    if not template.exists():
        print("  (DORA_DB.xlsx template not found, skipping Excel export)")
        return ""

    wb = openpyxl.load_workbook(template)

    # --- Helper: parse date string to datetime (midnight, no tz) for Excel ---
    def to_date(s: str | None):
        if not s:
            return None
        dt = parse_dt(s)
        if dt is None:
            return None
        return datetime(dt.year, dt.month, dt.day)

    # ── 1. Build Deployments rows ──
    if mode == "pipelines":
        # Source: CFR details (all completed builds)
        dep_source = cfr.get("_details", {})
    else:
        # Source: DF details (merged PRs)
        dep_source = df.get("_details", {})

    dep_rows = []
    for mk in sorted(dep_source.keys()):
        for rec in dep_source[mk]:
            date_str = rec.get("date", "")
            if mode == "pipelines":
                failed = 1 if rec.get("result", "") in ("failed", "partiallySucceeded") else 0
                source_id = rec.get("build_id")
            else:
                failed = 0
                source_id = rec.get("pr_id")
            dep_rows.append({
                "date": to_date(date_str),
                "failed": failed,
                "source_id": source_id,
            })
    dep_rows.sort(key=lambda r: r["date"] or datetime.min.date())

    # Assign sequential DeploymentIDs and build lookup map
    source_id_to_dep_id: dict[int | str, int] = {}
    for i, row in enumerate(dep_rows, 1):
        row["dep_id"] = i
        if row["source_id"] is not None:
            source_id_to_dep_id[row["source_id"]] = i

    # ── 2. Build Commits rows ──
    lt_details = lt.get("_details", {})
    commit_rows = []
    for mk in sorted(lt_details.keys()):
        for rec in lt_details[mk]:
            sid = rec.get("build_id") or rec.get("pr_id")
            commit_rows.append({
                "commit_id": sid,
                "date_commit": to_date(rec.get("commit_date")),
                "deployment_id": source_id_to_dep_id.get(sid, 0),
            })
    commit_rows.sort(key=lambda r: r["date_commit"] or datetime.min)

    # ── 3. Build Issues rows ──
    # Build a reverse lookup: date → list of DeploymentIDs for FK matching
    date_to_dep_ids: dict[datetime, list[int]] = {}
    for row in dep_rows:
        if row["date"]:
            date_to_dep_ids.setdefault(row["date"], []).append(row["dep_id"])

    mttr_details = mttr.get("_details", {})
    issue_rows = []
    for mk in sorted(mttr_details.keys()):
        for rec in mttr_details[mk]:
            recovery_date = to_date(rec.get("recovery_date"))
            # Match recovery date to a DeploymentID
            fixed_release_id = 0
            if recovery_date and recovery_date in date_to_dep_ids:
                fixed_release_id = date_to_dep_ids[recovery_date][0]
            issue_rows.append({
                "report_date": to_date(rec.get("failure_date")),
                "fixed_release_id": fixed_release_id,
            })
    issue_rows.sort(key=lambda r: r["report_date"] or datetime.min)

    # ── 4. Write Deployments tab ──
    ws_dep = wb["Deployments"]
    # Clear old data rows (keep header row 1)
    for row_idx in range(2, ws_dep.max_row + 1):
        for col_idx in range(1, 5):
            ws_dep.cell(row=row_idx, column=col_idx).value = None

    for i, row in enumerate(dep_rows, 2):
        ws_dep.cell(row=i, column=1).value = row["dep_id"]
        c_acc = ws_dep.cell(row=i, column=2)
        c_acc.value = row["date"]
        c_acc.number_format = r'dd\.mm\.yy;@'
        c_prod = ws_dep.cell(row=i, column=3)
        c_prod.value = row["date"]
        c_prod.number_format = 'd/m/yy;@'
        ws_dep.cell(row=i, column=4).value = row["failed"]

    # Resize table
    if dep_rows:
        last_row = 1 + len(dep_rows)
        ws_dep.tables["Deployments"].ref = f"A1:D{last_row}"

    # ── 5. Write Commits tab ──
    ws_com = wb["Commits"]
    for row_idx in range(2, ws_com.max_row + 1):
        for col_idx in range(1, 5):
            ws_com.cell(row=row_idx, column=col_idx).value = None

    COMMIT_LT_FORMULA = '=_xlfn.XLOOKUP(Commits[[#This Row],[DeploymentID]],Deployments[DeploymentID],Deployments[Prod])-Commits[[#This Row],[DateCommit]]'
    for i, row in enumerate(commit_rows, 2):
        ws_com.cell(row=i, column=1).value = row["commit_id"]
        c_date = ws_com.cell(row=i, column=2)
        c_date.value = row["date_commit"]
        c_date.number_format = 'dd/mm/yy;@'
        ws_com.cell(row=i, column=3).value = row["deployment_id"]
        c_lt = ws_com.cell(row=i, column=4)
        c_lt.value = COMMIT_LT_FORMULA
        c_lt.number_format = '0'

    if commit_rows:
        last_row = 1 + len(commit_rows)
        ws_com.tables["Commits"].ref = f"A1:D{last_row}"

    # ── 6. Write Issues tab ──
    ws_iss = wb["Issues"]
    for row_idx in range(2, ws_iss.max_row + 1):
        for col_idx in range(1, 6):
            ws_iss.cell(row=row_idx, column=col_idx).value = None

    ISSUE_LT_FORMULA = '=_xlfn.XLOOKUP(Issues[[#This Row],[Fixed-ReleaseID]],Deployments[DeploymentID],Deployments[Prod])-Issues[[#This Row],[Report Date]]'
    ISSUE_RD_FORMULA = '=_xlfn.XLOOKUP(Issues[[#This Row],[Fixed-ReleaseID]],Deployments[DeploymentID],Deployments[Prod])'
    for i, row in enumerate(issue_rows, 2):
        ws_iss.cell(row=i, column=1).value = i - 1  # sequential Issue ID
        c_report = ws_iss.cell(row=i, column=2)
        c_report.value = row["report_date"]
        c_report.number_format = r'dd\.mm\.yy;@'
        ws_iss.cell(row=i, column=3).value = row["fixed_release_id"]
        c_dtr = ws_iss.cell(row=i, column=4)
        c_dtr.value = ISSUE_LT_FORMULA
        c_dtr.number_format = '0.00'
        c_rd = ws_iss.cell(row=i, column=5)
        c_rd.value = ISSUE_RD_FORMULA
        c_rd.number_format = 'd/mm/yyyy;@'

    if issue_rows:
        last_row = 1 + len(issue_rows)
        ws_iss.tables["Issues"].ref = f"A1:E{last_row}"

    # ── 7. Update DORA tab month-start dates ──
    ws_dora = wb["DORA"]
    # months list is like ["2025-08", "2025-09", ...] — write as 1st of each month
    # DORA tab uses columns D–O (4–15) for up to 12 months
    for col_idx in range(4, 16):
        ws_dora.cell(row=1, column=col_idx).value = None
    for j, mk in enumerate(months):
        year, month = map(int, mk.split("-"))
        ws_dora.cell(row=1, column=4 + j).value = datetime(year, month, 1)

    # ── 8. Save ──
    reports_dir = Path(__file__).resolve().parent / "reports"
    reports_dir.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_project = project.replace(" ", "_")
    filepath = reports_dir / f"DORA_DB_{safe_project}_{mode}_{timestamp}.xlsx"
    wb.save(filepath)
    print(f"Excel report exported to: {filepath}")
    return str(filepath)


def print_detail_table(metric_abbr: str, label: str, month: str, records: list[dict]):
    """Print a formatted detail table for a single metric+month."""
    if not records:
        print(f"\n  No detail records for {label} in {month}.")
        return

    print(f"\n  {label} — {month}  ({len(records)} records)")
    print(f"  {'─' * 90}")

    if metric_abbr == "df":
        print(f"  {'Date':<28} {'Pipeline/Repo':<35} {'ID':>10}")
        print(f"  {'─' * 90}")
        for r in sorted(records, key=lambda x: x.get("date", "")):
            name = r.get("pipeline") or r.get("repo", "")
            rid = r.get("build_id") or r.get("pr_id", "")
            print(f"  {r.get('date', ''):<28} {name:<35} {str(rid):>10}")

    elif metric_abbr == "lt":
        print(f"  {'Commit Date':<28} {'Finish Date':<28} {'Lead Time':>12} {'Pipeline/Repo':<20} {'ID':>8}")
        print(f"  {'─' * 100}")
        for r in sorted(records, key=lambda x: x.get("commit_date", "")):
            name = r.get("pipeline") or r.get("repo", "")
            rid = r.get("build_id") or r.get("pr_id", "")
            lt_h = r.get("lead_time_hours")
            lt_str = format_hours(lt_h) if lt_h is not None else "N/A"
            print(f"  {r.get('commit_date', ''):<28} {r.get('finish_date', ''):<28} {lt_str:>12} {name:<20} {str(rid):>8}")

    elif metric_abbr == "cfr":
        print(f"  {'Date':<28} {'Pipeline/Repo':<35} {'Result':<20} {'Build ID':>10}")
        print(f"  {'─' * 95}")
        for r in sorted(records, key=lambda x: x.get("date", "")):
            result = r.get("result", "")
            marker = " *** FAILED" if result in ("failed", "partiallySucceeded") else ""
            print(f"  {r.get('date', ''):<28} {r.get('pipeline', ''):<35} {result + marker:<20} {str(r.get('build_id', '')):>10}")

    elif metric_abbr == "mttr":
        print(f"  {'Failure Date':<28} {'Recovery Date':<28} {'Duration':>14} {'Pipeline':<20}")
        print(f"  {'─' * 92}")
        for r in sorted(records, key=lambda x: x.get("failure_date", "")):
            dur = format_hours(r.get("duration_hours"))
            print(f"  {r.get('failure_date', ''):<28} {r.get('recovery_date', ''):<28} {dur:>14} {r.get('pipeline', ''):<20}")


def drill_down(df: dict, lt: dict, cfr: dict, mttr: dict, months: list[str]):
    """Interactive drill-down loop after displaying results."""
    metric_map = {
        "df": ("Deployment Frequency", df.get("_details", {})),
        "lt": ("Lead Time", lt.get("_details", {})),
        "cfr": ("Change Failure Rate", cfr.get("_details", {})),
        "mttr": ("MTTR", mttr.get("_details", {})),
    }

    while True:
        try:
            raw = input('\nDrill into a cell? (e.g. "df 2026-01", or "q" to quit): ').strip()
        except (EOFError, KeyboardInterrupt):
            break
        if not raw or raw.lower() == "q":
            break

        parts = raw.split(None, 1)
        if len(parts) != 2:
            print("  Format: <metric> <month>  (e.g. df 2026-01)")
            print(f"  Metrics: {', '.join(metric_map.keys())}")
            continue

        abbr, month = parts[0].lower(), parts[1].strip()
        if abbr not in metric_map:
            print(f"  Unknown metric '{abbr}'. Use: {', '.join(metric_map.keys())}")
            continue
        if month not in months:
            print(f"  Month '{month}' not in range. Available: {', '.join(months)}")
            continue

        label, details_dict = metric_map[abbr]
        records = details_dict.get(month, [])
        print_detail_table(abbr, label, month, records)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="DORA Metrics CLI — Azure DevOps")
    parser.add_argument("-org", dest="org", help="Azure DevOps organization (name or URL)")
    parser.add_argument("-project", dest="project", help="Project name, or 'all' for all projects")
    parser.add_argument("-mode", dest="mode", choices=["pipelines", "pullrequests"], help="Measure using pipelines or pull requests")
    return parser.parse_args()


async def main():
    args = parse_args()

    print("=" * 50)
    print("  DORA Metrics CLI — Azure DevOps")
    print("=" * 50)

    pat = read_pat()
    print("\nPAT loaded from env.tks")

    org_input = args.org
    if not org_input:
        org_input = input("\nEnter Azure DevOps organization (name or URL): ").strip()
    if not org_input:
        print("No organization provided.")
        sys.exit(1)
    org = parse_org(org_input)
    print(f"Organization: {org}")

    async with httpx.AsyncClient() as client:
        # Fetch projects
        print("\nFetching projects...")
        try:
            projects = await fetch_projects(client, org, pat)
        except httpx.HTTPStatusError as e:
            print(f"Error fetching projects: {e.response.status_code} {e.response.text[:200]}")
            sys.exit(1)

        if not projects:
            print("No projects found.")
            sys.exit(1)

        print(f"\nFound {len(projects)} project(s):")
        project_names = [p["name"] for p in projects]

        if args.project and args.project.lower() == "all":
            selected_projects = projects
        if args.project and args.project.lower() != "all":
            matching = [p for p in projects if p["name"].lower() == args.project.lower()]
            if not matching:
                print(f"Project '{args.project}' not found. Available: {', '.join(project_names)}")
                sys.exit(1)
            selected_projects = matching
        if not args.project:
            proj_choice = prompt_choice(project_names, "a project (0 for all)", allow_all=True)
            selected_projects = projects if proj_choice is None else [projects[proj_choice]]
        print(f"\nSelected project(s): {', '.join(p['name'] for p in selected_projects)}")

        # Mode selection
        mode = args.mode
        if not mode:
            print("\nMeasure DORA metrics using:")
            mode_choice = prompt_choice(["Pipelines", "Pull Requests"], "a mode")
            mode = "pipelines" if mode_choice == 0 else "pullrequests"
        print(f"Mode: {mode}")

        run_per_project = len(selected_projects) > 1

        # ── Pipeline mode ──
        if mode == "pipelines":
            project_builds: dict[str, list[dict]] = {}
            for proj in selected_projects:
                pname = proj["name"]
                print(f"\nFetching pipelines for '{pname}'...")
                pipelines = await fetch_pipelines(client, org, pname, pat)
                if not pipelines:
                    print(f"  No pipelines found in '{pname}', skipping.")
                    continue

                if not run_per_project:
                    print(f"\nFound {len(pipelines)} pipeline(s):")
                    choice = prompt_choice([p["name"] for p in pipelines], "a pipeline (0 for all)", allow_all=True)
                    selected_pipelines = pipelines if choice is None else [pipelines[choice]]
                    print(f"\nSelected: {', '.join(p['name'] for p in selected_pipelines)}")
                if run_per_project:
                    selected_pipelines = pipelines
                    print(f"  Found {len(pipelines)} pipeline(s), fetching all...")

                proj_builds: list[dict] = []
                for p in selected_pipelines:
                    print(f"  Fetching builds for '{pname}/{p['name']}'...")
                    builds = await fetch_builds(client, org, pname, p["id"], pat)
                    print(f"    → {len(builds)} builds")
                    proj_builds.extend(builds)
                if proj_builds:
                    project_builds[pname] = proj_builds

            if not project_builds:
                print("\nNo builds found in the past 6 months.")
                sys.exit(1)

            if run_per_project:
                for pname, builds in project_builds.items():
                    print(f"\n{'#' * 40}")
                    print(f"  Computing metrics for '{pname}' ({len(builds)} builds)...")
                    months = all_months_in_range(builds)
                    df = compute_deployment_frequency(builds)
                    cfr = compute_change_failure_rate_by_month(builds)
                    mttr_result = compute_mttr_by_month(builds)
                    print(f"  Fetching commit data for lead time...")
                    lt = await compute_lead_times_by_month(client, org, builds, pat)
                    print_results(df, lt, cfr, mttr_result, months, title=f"DORA METRICS — {pname} [Pipelines]")
                    export_json(org, pname, "pipelines", months, df, lt, cfr, mttr_result)
                    export_excel(org, pname, "pipelines", months, df, lt, cfr, mttr_result)
                    drill_down(df, lt, cfr, mttr_result, months)

            if not run_per_project:
                all_builds = list(project_builds.values())[0]
                pname = list(project_builds.keys())[0]
                print(f"\nTotal builds: {len(all_builds)}")
                print("\nComputing metrics...")
                months = all_months_in_range(all_builds)
                df = compute_deployment_frequency(all_builds)
                cfr = compute_change_failure_rate_by_month(all_builds)
                mttr_result = compute_mttr_by_month(all_builds)
                print("Fetching commit data for lead time (this may take a moment)...")
                lt = await compute_lead_times_by_month(client, org, all_builds, pat)
                print_results(df, lt, cfr, mttr_result, months, title=f"DORA METRICS — {pname} [Pipelines]")
                export_json(org, pname, "pipelines", months, df, lt, cfr, mttr_result)
                export_excel(org, pname, "pipelines", months, df, lt, cfr, mttr_result)
                drill_down(df, lt, cfr, mttr_result, months)

        # ── Pull Request mode ──
        if mode == "pullrequests":
            for proj in selected_projects:
                pname = proj["name"]
                print(f"\nFetching repos for '{pname}'...")
                repos = await fetch_repos(client, org, pname, pat)
                if not repos:
                    print(f"  No repos found in '{pname}', skipping.")
                    continue

                # Fetch PRs from all repos
                all_prs: list[dict] = []
                for repo in repos:
                    print(f"  Fetching PRs for repo '{repo['name']}'...")
                    completed = await fetch_pull_requests(client, org, pname, repo["id"], pat, status="completed")
                    abandoned = await fetch_pull_requests(client, org, pname, repo["id"], pat, status="abandoned")
                    repo_prs = completed + abandoned
                    if repo_prs:
                        print(f"    → {len(completed)} merged, {len(abandoned)} abandoned")
                    all_prs.extend(repo_prs)

                if not all_prs:
                    print(f"  No PRs found in '{pname}' in the past 6 months, skipping.")
                    continue

                merged_count = len([pr for pr in all_prs if pr.get("status") == "completed"])
                abandoned_count = len([pr for pr in all_prs if pr.get("status") == "abandoned"])
                print(f"\n  Total PRs for '{pname}': {len(all_prs)} ({merged_count} merged, {abandoned_count} abandoned)")

                # Fetch builds for CFR/MTTR
                print(f"  Fetching builds for CFR/MTTR...")
                all_builds = await fetch_all_builds_for_project(client, org, pname, pat)
                print(f"    → {len(all_builds)} builds")

                # Compute PR-mode metrics
                print(f"  Computing metrics...")
                months = all_months_in_range(all_prs)
                df = compute_pr_deployment_frequency(all_prs)
                cfr = compute_change_failure_rate_by_month(all_builds)
                mttr_result = compute_mttr_by_month(all_builds)

                print(f"  Fetching commit data for lead time (this may take a moment)...")
                lt = await compute_pr_lead_times_by_month(client, org, pname, all_prs, pat)

                print_results(df, lt, cfr, mttr_result, months, title=f"DORA METRICS — {pname} [Pull Requests]")
                export_json(org, pname, "pullrequests", months, df, lt, cfr, mttr_result)
                export_excel(org, pname, "pullrequests", months, df, lt, cfr, mttr_result)
                drill_down(df, lt, cfr, mttr_result, months)


if __name__ == "__main__":
    asyncio.run(main())
