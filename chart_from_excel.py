#!/usr/bin/env python3
"""Generate DORA charts from a manually populated Excel workbook.

Reads any sheet whose name ends with '_Manual', extracts the DORA metrics
from the expected row layout, and produces a branded PNG chart via dora_charts.

Row layout (data columns start at F, column 6):
  Row 1 — Date of measurement
  Row 2 — Releases to ACC since last measurement
  Row 3 — Releases to PROD since last measurement
  Row 4 — Failed releases to PROD (count) OR CFR as a decimal fraction (0.18 = 18%)
  Row 5 — Average lead time: days from commit to production
  Row 6 — Average MTTR: days to recover from important failures

Usage:
    python chart_from_excel.py <path-to-excel>
    python chart_from_excel.py reports/DORA_DB_Generator.xlsx
"""

import calendar
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))

from dora_charts import generate_charts

# First column that contains temporal data (1-based, F = 6)
DATA_START_COL = 6

# Row indices (1-based) for each metric
ROW_DATE = 1
ROW_ACC_RELEASES = 2
ROW_PROD_RELEASES = 3
ROW_CFR = 4
ROW_LEAD_TIME = 5
ROW_MTTR = 6


def _read_sheet(ws) -> tuple[list[str], dict, dict, dict, dict]:
    """Parse a _Manual worksheet and return (months, df, lt, cfr, mttr) dicts."""

    dates: list[datetime] = []
    acc_rel: list[float] = []
    prod_rel: list[float] = []
    cfr_raw: list[float] = []
    lt_days: list[float] = []
    mttr_days_list: list[float] = []

    for col in range(DATA_START_COL, ws.max_column + 1):
        date_val = ws.cell(row=ROW_DATE, column=col).value
        if date_val is None:
            break
        if not isinstance(date_val, datetime):
            continue
        dates.append(date_val)
        acc_rel.append(ws.cell(row=ROW_ACC_RELEASES, column=col).value or 0)
        prod_rel.append(ws.cell(row=ROW_PROD_RELEASES, column=col).value or 0)
        cfr_raw.append(ws.cell(row=ROW_CFR, column=col).value or 0)
        lt_days.append(ws.cell(row=ROW_LEAD_TIME, column=col).value or 0)
        mttr_days_list.append(ws.cell(row=ROW_MTTR, column=col).value or 0)

    if not dates:
        raise ValueError("No date row found — check that row 1 contains dates from column F onwards")

    # Drop future dates (dates after today)
    today = datetime.now()
    dates_and_data = [
        (d, acc_rel[i], prod_rel[i], cfr_raw[i], lt_days[i], mttr_days_list[i])
        for i, d in enumerate(dates)
        if d <= today
    ]
    if not dates_and_data:
        raise ValueError("All dates are in the future — nothing to chart")

    # Restrict to the last 6 non-future months
    dates_and_data = dates_and_data[-6:]
    dates, acc_rel, prod_rel, cfr_raw, lt_days, mttr_days_list = (
        list(col) for col in zip(*dates_and_data)
    )

    # Keep only months where at least one metric is non-zero
    active = [
        i for i, d in enumerate(dates)
        if any([acc_rel[i], prod_rel[i], cfr_raw[i], lt_days[i], mttr_days_list[i]])
    ]
    if not active:
        raise ValueError("All metric values are zero — nothing to chart")

    months = [dates[i].strftime("%Y-%m") for i in active]

    # ── Deployment Frequency ────────────────────────────────────────────────────
    # days_per_dep = days_in_measurement_period / prod_releases
    # Period length: gap to the previous active measurement, or days-in-month for the first.
    df_monthly: dict = {}
    df_vals: list[float] = []
    for idx, i in enumerate(active):
        d = dates[i]
        if idx > 0:
            prev_d = dates[active[idx - 1]]
            period_days = (d - prev_d).days
        else:
            period_days = calendar.monthrange(d.year, d.month)[1]

        mk = d.strftime("%Y-%m")
        n = prod_rel[i]
        if n and n > 0:
            dpd = period_days / n
            df_monthly[mk] = {"days_per_dep": dpd}
            df_vals.append(dpd)
        else:
            df_monthly[mk] = {"days_per_dep": None}

    df = {
        "monthly": df_monthly,
        "overall_days_per_dep": sum(df_vals) / len(df_vals) if df_vals else None,
    }

    # ── Lead Time ───────────────────────────────────────────────────────────────
    # Source values are in days; chart expects hours.
    lt: dict = {}
    lt_vals: list[float] = []
    for i in active:
        mk = dates[i].strftime("%Y-%m")
        v = lt_days[i]
        if v and v > 0:
            hours = v * 24
            lt[mk] = {"avg_hours": hours}
            lt_vals.append(hours)
        else:
            lt[mk] = {"avg_hours": None}
    lt["_overall"] = {"avg_hours": sum(lt_vals) / len(lt_vals) if lt_vals else None}

    # ── Change Failure Rate ─────────────────────────────────────────────────────
    # Accepts either:
    #   • decimal fraction  (0 < v ≤ 1)  → multiply by 100 for %
    #   • failure count     (v > 1)       → divide by prod_releases × 100 for %
    cfr: dict = {}
    cfr_vals: list[float] = []
    for idx, i in enumerate(active):
        mk = dates[i].strftime("%Y-%m")
        v = cfr_raw[i]
        n = prod_rel[i]
        if not v or v == 0:
            cfr[mk] = {"rate_pct": None}
        elif 0 < v <= 1.0:
            pct = v * 100
            cfr[mk] = {"rate_pct": pct}
            cfr_vals.append(pct)
        else:
            # Treat as failure count
            if n and n > 0:
                pct = (v / n) * 100
                cfr[mk] = {"rate_pct": pct}
                cfr_vals.append(pct)
            else:
                cfr[mk] = {"rate_pct": None}
    cfr["_overall"] = {"rate_pct": sum(cfr_vals) / len(cfr_vals) if cfr_vals else None}

    # ── Mean Time to Recovery ────────────────────────────────────────────────────
    # Source values are in days; chart expects hours.
    mttr: dict = {}
    mttr_vals: list[float] = []
    for i in active:
        mk = dates[i].strftime("%Y-%m")
        v = mttr_days_list[i]
        if v and v > 0:
            hours = v * 24
            mttr[mk] = {"avg_hours": hours}
            mttr_vals.append(hours)
        else:
            mttr[mk] = {"avg_hours": None}
    mttr["_overall"] = {"avg_hours": sum(mttr_vals) / len(mttr_vals) if mttr_vals else None}

    return months, df, lt, cfr, mttr


def generate_from_excel(excel_path: str | Path) -> list[str]:
    """Process all _Manual sheets in the workbook and return list of PNG paths."""
    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    manual_sheets = [name for name in wb.sheetnames if name.endswith("_Manual")]
    if not manual_sheets:
        raise ValueError(f"No sheet ending with '_Manual' found in {excel_path.name}")

    outputs = []
    for sheet_name in manual_sheets:
        team = sheet_name[: -len("_Manual")]
        ws = wb[sheet_name]
        print(f"\nProcessing sheet '{sheet_name}' → team: '{team}'")

        months, df, lt, cfr, mttr = _read_sheet(ws)
        print(f"  Months with data: {months[0]} – {months[-1]} ({len(months)} points)")

        out = generate_charts(team, "manual", months, df, lt, cfr, mttr)
        outputs.append(out)

    return outputs


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    paths = generate_from_excel(sys.argv[1])
    print(f"\nDone. {len(paths)} chart(s) generated.")
